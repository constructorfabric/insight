"""Read back what the browser downloaded, and what the grid behind it showed.

A download event and a byte count prove the transport worked, which was never
the part in doubt. What a journey has to prove is that the file holds the table
the user was looking at — and CSV and XLSX are written by different code
(client-side for timeseries exports, `rust_xlsxwriter` beside the `csv` crate
for evidence exports), so the two formats are read back and compared against
each other as well as against the grid.

Every cell — from a CSV string, an XLSX number, or a rendered DOM node — is put
through one normalizer, so `29`, `29.0` and a formatted `"+2,705"` compare as
the same value while an empty cell stays distinct from a zero. That distinction
is the whole point: a serializer that turns a missing value into 0 passes any
assertion that only counts rows.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Locator, Page

#: A table read out of an export or off the screen: normalized cells, row-wise,
#: header rows included.
Table = list[list[str]]

#: What the SPA renders in a cell that has no value — em dash, en dash or
#: hyphen. Read as empty, so a rendered row compares against its export.
_EMPTY_MARKERS = frozenset({"\u2014", "\u2013", "-", ""})

#: Digit-group separators and signs the UI inserts and no exporter emits:
#: narrow no-break space, no-break space, comma, plus.
_GROUPING = ("\u202f", "\u00a0", ",", "+")


def download_export(
    page: Page, menu_item: str, *, into: Path, exact: bool = True
) -> tuple[str, Table]:
    """Click an export menu item, save the file, and read it back as a table."""
    with page.expect_download() as download_info:
        page.get_by_role("menuitem", name=menu_item, exact=exact).click()

    download = download_info.value
    destination = into / download.suggested_filename
    download.save_as(destination)

    return download.suggested_filename, read_export(destination)


def read_export(path: Path) -> Table:
    if path.suffix == ".csv":
        return _read_csv(path)
    if path.suffix == ".xlsx":
        return _read_xlsx(path)
    raise AssertionError(f"no reader for {path.name}; the suite reads .csv and .xlsx")


def rendered_rows(table: Locator) -> Table:
    """The grid as the browser rendered it, normalized like an exported table.

    A virtualized grid keeps only its visible window in the DOM, so this is the
    window, not the whole result — compare it against the head of an export and
    take the row total from `aria-rowcount`.
    """
    rows: Table = []
    for row in table.get_by_role("row").all():
        cells = row.locator('[role="cell"], [role="columnheader"], td, th')
        rows.append([_cell(text) for text in cells.all_inner_texts()])

    return rows


def claimed_row_count(table: Locator) -> int:
    """What the grid says its full result set holds, virtualization aside."""
    declared = table.get_attribute("aria-rowcount")
    assert declared is not None, "grid declares no aria-rowcount to reconcile the export against"

    return int(declared)


def data_rows(table: Table, *, after: int) -> Table:
    """Everything below the header rows, blank padding dropped."""
    return [row for row in table[after:] if any(cell for cell in row)]


def _read_csv(path: Path) -> Table:
    # The exporters lead with a BOM so a spreadsheet keeps non-ASCII names.
    text = path.read_text(encoding="utf-8-sig")

    return [[_cell(value) for value in row] for row in csv.reader(text.splitlines()) if any(row)]


def _read_xlsx(path: Path) -> Table:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = [[_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return [row for row in rows if any(row)]


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, int | float):
        return _number(float(value))

    text = " ".join(str(value).split())
    if text in _EMPTY_MARKERS:
        return ""

    plain = text.replace("\u2212", "-")  # the UI signs negatives with U+2212
    for separator in _GROUPING:
        plain = plain.replace(separator, "")
    try:
        return _number(float(plain))
    except ValueError:
        return text


def _number(value: float) -> str:
    return str(int(value)) if value.is_integer() else repr(value)
